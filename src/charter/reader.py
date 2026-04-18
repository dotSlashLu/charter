"""Read an Excel workbook and extract a schema summary for the AI prompt."""

from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import log

_log = log.get("reader")

SAMPLE_ROWS = 5
TYPE_PROBE_ROWS = 20


@dataclass
class ColumnInfo:
    name: str
    dtype: str  # "numeric", "text", "date", "boolean", "empty"
    index: int  # 1-based column index in the sheet


@dataclass
class SheetSchema:
    sheet_name: str
    columns: list[ColumnInfo] = field(default_factory=list)
    sample_rows: list[list[object]] = field(default_factory=list)
    total_rows: int = 0


def _infer_dtype(values: list[object]) -> str:
    dominated: dict[str, int] = {}
    for v in values:
        if v is None:
            continue
        if isinstance(v, bool):
            dominated["boolean"] = dominated.get("boolean", 0) + 1
        elif isinstance(v, (int, float)):
            dominated["numeric"] = dominated.get("numeric", 0) + 1
        elif isinstance(v, datetime.datetime | datetime.date):
            dominated["date"] = dominated.get("date", 0) + 1
        else:
            dominated["text"] = dominated.get("text", 0) + 1
    if not dominated:
        return "empty"
    return max(dominated, key=lambda k: dominated[k])


def read_schema(path: str | Path) -> tuple[SheetSchema, Worksheet]:
    """Return the schema of the first worksheet and the worksheet itself."""
    _log.debug("Loading workbook: %s", path)
    wb = load_workbook(str(path))
    ws = wb.active
    assert ws is not None, "Workbook has no active sheet"
    _log.debug("Active sheet: '%s'  max_row=%s  max_column=%s", ws.title, ws.max_row, ws.max_column)

    # Read max_row before any iter_rows call (openpyxl may inflate it).
    total_rows = ws.max_row - 1 if ws.max_row else 0  # exclude header

    headers: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value) if cell.value is not None else f"Column{cell.column}")
    _log.debug("Headers (%d columns): %s", len(headers), headers)

    probe_end = min(1 + TYPE_PROBE_ROWS, 1 + total_rows)
    _log.debug("Probing rows 2..%d for type inference", probe_end)
    data_rows: list[list[object]] = []
    for row in ws.iter_rows(min_row=2, max_row=probe_end, values_only=True):
        data_rows.append(list(row))

    columns: list[ColumnInfo] = []
    for col_idx, name in enumerate(headers):
        col_values = [r[col_idx] for r in data_rows if col_idx < len(r)]
        dtype = _infer_dtype(col_values)
        columns.append(ColumnInfo(name=name, dtype=dtype, index=col_idx + 1))
        _log.debug("  Column %d '%s' -> %s", col_idx + 1, name, dtype)

    sample = data_rows[:SAMPLE_ROWS]
    _log.debug("Total data rows: %d  sample rows: %d", total_rows, len(sample))

    schema = SheetSchema(
        sheet_name=ws.title or "Sheet1",
        columns=columns,
        sample_rows=sample,
        total_rows=total_rows,
    )
    return schema, ws


def schema_to_prompt(schema: SheetSchema) -> str:
    """Format a SheetSchema into a human-readable string for the LLM prompt."""
    col_desc = ", ".join(f"{c.name}({c.dtype})" for c in schema.columns)
    lines = [
        f"Sheet: {schema.sheet_name}",
        f"Total data rows: {schema.total_rows}",
        f"Columns: [{col_desc}]",
        "",
        "Sample data (first 5 rows):",
    ]
    header_line = " | ".join(c.name for c in schema.columns)
    lines.append(header_line)
    lines.append("-" * len(header_line))
    for row in schema.sample_rows:
        lines.append(" | ".join(_fmt(v) for v in row))
    return "\n".join(lines)


def _fmt(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:g}"
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, datetime.date):
        return v.isoformat()
    return str(v)
