"""Convert a DSL pipeline into Excel sheets populated with formulas."""

from __future__ import annotations

import datetime

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import log
from .dsl import (
    AggregationValue,
    AnyAction,
    DSLSpec,
    FilterAction,
    GroupByAction,
    PivotAction,
    SelectAction,
    SortAction,
)
from .reader import SheetSchema

_log = log.get("executor")

_pipe_counter = 0


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def execute_dsl(
    spec: DSLSpec,
    schema: SheetSchema,
    ws_source: Worksheet,
    wb: "Workbook",  # noqa: F821 – forward ref to openpyxl
    sheet_title: str,
) -> Worksheet:
    """Execute a DSL pipeline and write the final result to a new sheet.

    When a ``filter`` step precedes a ``groupby`` or ``pivot``, the filter
    conditions are folded directly into SUMIFS/COUNTIFS criteria instead of
    creating an unreliable intermediate formula sheet.  Other step
    combinations still materialise helper sheets as needed.
    """
    global _pipe_counter

    steps = spec.pipeline
    _log.debug(
        "execute_dsl: %d step(s)  sheet='%s'  actions=[%s]",
        len(steps),
        sheet_title,
        ", ".join(s.action for s in steps),
    )

    current_source = _source_info_from_schema(schema, ws_source)
    _log.debug(
        "  Initial source: sheet='%s'  columns=%s  data_rows=%d",
        current_source.sheet_name,
        current_source.columns,
        current_source.data_rows,
    )

    pending_filters: list[FilterAction] = []
    last_ws: Worksheet | None = None

    for i, step in enumerate(steps):
        is_last = i == len(steps) - 1

        # --- Accumulate AND-logic filters (will be folded into groupby/pivot) ---
        if isinstance(step, FilterAction) and step.logic == "AND" and not is_last:
            for cond in step.conditions:
                current_source._check_col(cond.column)
            pending_filters.append(step)
            _log.debug(
                "  Step %d/%d: action=filter (accumulated, %d conditions)",
                i + 1, len(steps), len(step.conditions),
            )
            continue

        # --- If we have pending filters and the step can't fold them,
        #     materialise intermediate filter sheets first ---
        if pending_filters and not isinstance(step, (GroupByAction, PivotAction)):
            current_source = _materialise_filters(
                pending_filters, current_source, ws_source, schema, wb,
            )
            pending_filters = []

        # --- Create output sheet ---
        if is_last:
            target_title = sheet_title[:31]
        else:
            _pipe_counter += 1
            target_title = f"_pipe_{_pipe_counter}_{step.action}"[:31]

        _log.debug(
            "  Step %d/%d: action=%s -> sheet='%s'%s",
            i + 1, len(steps), step.action, target_title,
            f"  (folding {sum(len(f.conditions) for f in pending_filters)} filter conditions)"
            if pending_filters else "",
        )

        ws_out = wb.create_sheet(title=target_title)

        match step:
            case GroupByAction():
                _write_groupby(step, current_source, ws_out, pending_filters)
                pending_filters = []
            case PivotAction():
                _write_pivot(step, current_source, ws_out, pending_filters)
                pending_filters = []
            case FilterAction():
                _write_filter(step, current_source, ws_out)
            case SelectAction():
                _write_select(step, current_source, ws_out)
            case SortAction():
                _write_sort(step, current_source, ws_out)

        # Propagate number formats from source to output.
        if current_source.col_formats:
            header_map: dict[str, int] = {}
            for cell in next(ws_out.iter_rows(min_row=1, max_row=1)):
                if cell.value is not None:
                    header_map[str(cell.value)] = cell.column
            _apply_formats(ws_out, header_map, current_source.col_formats, ws_out.max_row)

        out_rows = ws_out.max_row - 1 if ws_out.max_row else 0
        _log.debug("    -> %d data rows written", out_rows)

        current_source = _source_info_from_ws(ws_out, ws_source, schema)
        last_ws = ws_out

    assert last_ws is not None
    return last_ws


def _materialise_filters(
    filters: list[FilterAction],
    current_source: _SourceInfo,
    ws_source: Worksheet,
    schema: SheetSchema,
    wb: "Workbook",
) -> _SourceInfo:
    """Create intermediate filter sheets when we can't fold into groupby/pivot."""
    global _pipe_counter
    for filt in filters:
        _pipe_counter += 1
        title = f"_pipe_{_pipe_counter}_filter"[:31]
        _log.debug("    Materialising filter -> '%s'", title)
        ws = wb.create_sheet(title=title)
        _write_filter(filt, current_source, ws)
        current_source = _source_info_from_ws(ws, ws_source, schema)
    return current_source


# ---------------------------------------------------------------------------
# Source resolution helpers
# ---------------------------------------------------------------------------

class _SourceInfo:
    """Describes where source data lives (sheet name, columns, row range)."""

    def __init__(
        self,
        sheet_name: str,
        columns: list[str],
        col_indices: dict[str, int],
        data_rows: int,
        ws: Worksheet | None = None,
        col_formats: dict[str, str] | None = None,
        *,
        data_ws: Worksheet | None = None,
        data_col_indices: dict[str, int] | None = None,
    ):
        self.sheet_name = sheet_name
        self.columns = columns
        self.col_indices = col_indices  # column name -> 1-based index
        self.data_rows = data_rows      # number of data rows (excl header)
        self.ws = ws                    # worksheet for formula references
        self.col_formats = col_formats or {}
        self.data_ws = data_ws or ws
        self.data_col_indices = data_col_indices or col_indices

    def col_letter(self, col_name: str) -> str:
        self._check_col(col_name)
        return get_column_letter(self.col_indices[col_name])

    def _check_col(self, col_name: str) -> None:
        if col_name not in self.col_indices:
            available = ", ".join(self.columns)
            raise KeyError(
                f"Column '{col_name}' not found in source sheet '{self.sheet_name}'. "
                f"Available columns: [{available}]. "
                f"If using a pipeline, each step can only reference columns "
                f"produced by the previous step (or the original table for step 1)."
            )

    def range(self, col_name: str) -> str:
        """Absolute data range like ``Sheet1!A$2:A$1500``."""
        letter = self.col_letter(col_name)
        end = self.data_rows + 1  # +1 because row 1 is header
        return f"'{self.sheet_name}'!{letter}$2:{letter}${end}"

    def cell(self, col_name: str, data_row: int) -> str:
        """Absolute cell ref for a specific data row (1-based among data)."""
        letter = self.col_letter(col_name)
        return f"'{self.sheet_name}'!{letter}{data_row + 1}"


def _source_info_from_schema(schema: SheetSchema, ws: Worksheet | None = None) -> _SourceInfo:
    cols = [c.name for c in schema.columns]
    col_map = {c.name: c.index for c in schema.columns}
    fmt_map: dict[str, str] = {}
    if ws is not None:
        fmt_map = _read_formats(ws, cols, col_map)
    return _SourceInfo(schema.sheet_name, cols, col_map, schema.total_rows, ws, fmt_map)


def _source_info_from_ws(
    ws_out: Worksheet,
    original_ws: Worksheet,
    schema: SheetSchema,
) -> _SourceInfo:
    """Build a _SourceInfo from a worksheet that was just written (may contain formulas)."""
    cols: list[str] = []
    col_map: dict[str, int] = {}
    for cell in next(ws_out.iter_rows(min_row=1, max_row=1)):
        name = str(cell.value)
        cols.append(name)
        col_map[name] = cell.column

    fmt_map = _read_formats(ws_out, cols, col_map)
    data_rows = max(ws_out.max_row - 1, 0)

    orig_col_map: dict[str, int] = {}
    for c in schema.columns:
        if c.name in col_map:
            orig_col_map[c.name] = c.index

    return _SourceInfo(
        ws_out.title, cols, col_map, data_rows, ws_out, fmt_map,
        data_ws=original_ws, data_col_indices=orig_col_map,
    )


def _read_formats(ws: Worksheet, columns: list[str], col_indices: dict[str, int]) -> dict[str, str]:
    """Read number_format from the first data row of *ws* for each column."""
    fmt: dict[str, str] = {}
    for name in columns:
        idx = col_indices[name]
        cell = ws.cell(row=2, column=idx)
        nf = cell.number_format
        if nf and nf != "General":
            fmt[name] = nf
    return fmt


def _apply_formats(ws: Worksheet, col_name_to_out_col: dict[str, int], formats: dict[str, str], max_row: int) -> None:
    """Apply source number formats to output columns."""
    for col_name, out_col in col_name_to_out_col.items():
        nf = formats.get(col_name)
        if not nf:
            continue
        for row in range(2, max_row + 1):
            ws.cell(row=row, column=out_col).number_format = nf


# ---------------------------------------------------------------------------
# Filter-folding helpers
# ---------------------------------------------------------------------------

def _build_sumifs_criteria(
    filters: list[FilterAction],
    src: _SourceInfo,
) -> list[str]:
    """Convert accumulated filter conditions into SUMIFS/COUNTIFS criteria pairs.

    Each pair is a string ``"range,criterion"`` ready to be appended to
    SUMIFS / COUNTIFS argument lists.
    """
    pairs: list[str] = []
    for filt in filters:
        for cond in filt.conditions:
            crit_range = src.range(cond.column)
            val = cond.value
            op = cond.operator

            if op == "=":
                crit_val = _formula_literal(val)
            elif op == "!=":
                inner = val if isinstance(val, str) else str(val)
                crit_val = f'"<>{inner}"'
            else:
                inner = val if isinstance(val, str) else str(val)
                crit_val = f'"{op}{inner}"'

            pairs.append(f"{crit_range},{crit_val}")
    return pairs


def _build_row_checks(
    filters: list[FilterAction],
    col_indices: dict[str, int],
) -> list[tuple[int, str, object]]:
    """Build (col_index, operator, value) tuples for Python-side row filtering."""
    checks: list[tuple[int, str, object]] = []
    for filt in filters:
        for cond in filt.conditions:
            checks.append((col_indices[cond.column], cond.operator, cond.value))
    return checks


def _row_passes(row: tuple, checks: list[tuple[int, str, object]]) -> bool:
    """Return True if *row* satisfies all filter checks."""
    for col_idx, op, target in checks:
        cell_val = row[col_idx - 1].value
        if not _compare_op(cell_val, op, target):
            return False
    return True


def _compare_op(cell_val: object, op: str, target: object) -> bool:
    if cell_val is None:
        return False
    if op == "=":
        return cell_val == target or str(cell_val) == str(target)
    if op == "!=":
        return cell_val != target and str(cell_val) != str(target)
    try:
        a = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
        b = float(target) if not isinstance(target, (int, float)) else target
    except (ValueError, TypeError):
        return False
    if op == ">":
        return a > b
    if op == "<":
        return a < b
    if op == ">=":
        return a >= b
    if op == "<=":
        return a <= b
    return False


# ---------------------------------------------------------------------------
# Action writers
# ---------------------------------------------------------------------------

def _write_select(dsl: SelectAction, src: _SourceInfo, ws: Worksheet) -> None:
    _log.debug("  [select] columns=%s  distinct=%s", dsl.columns, dsl.distinct)
    for out_col_idx, col_name in enumerate(dsl.columns, start=1):
        out_letter = get_column_letter(out_col_idx)
        ws.cell(row=1, column=out_col_idx, value=col_name)

        if dsl.distinct:
            src_range = src.range(col_name)
            for data_row in range(1, src.data_rows + 1):
                excel_row = data_row + 1
                prev = f"${out_letter}$1:{out_letter}{excel_row - 1}"
                formula = (
                    f'=IFERROR(INDEX({src_range},'
                    f'AGGREGATE(15,6,'
                    f'ROW({src_range})-ROW(INDIRECT("'
                    f"'{src.sheet_name}'!{src.col_letter(col_name)}$2\"))+1"
                    f'/(COUNTIF({prev},{src_range})=0)'
                    f',{data_row})),"")'
                )
                ws.cell(row=excel_row, column=out_col_idx, value=formula)
            _log.debug("    '%s': %d distinct-formula rows", col_name, src.data_rows)
        else:
            for data_row in range(1, src.data_rows + 1):
                ref = src.cell(col_name, data_row)
                ws.cell(row=data_row + 1, column=out_col_idx, value=f"={ref}")
            _log.debug("    '%s': %d ref rows", col_name, src.data_rows)


def _write_filter(dsl: FilterAction, src: _SourceInfo, ws: Worksheet) -> None:
    _log.debug(
        "  [filter] conditions=%s  logic=%s",
        [(c.column, c.operator, c.value) for c in dsl.conditions],
        dsl.logic,
    )
    all_cols = src.columns
    for out_col_idx, col_name in enumerate(all_cols, start=1):
        ws.cell(row=1, column=out_col_idx, value=col_name)

    cond_parts: list[str] = []
    for cond in dsl.conditions:
        src_range = src.range(cond.column)
        val = f'"{cond.value}"' if isinstance(cond.value, str) else str(cond.value)
        op = cond.operator
        if op == "=":
            cond_parts.append(f"({src_range}={val})")
        elif op == "!=":
            cond_parts.append(f"({src_range}<>{val})")
        else:
            cond_parts.append(f"({src_range}{op}{val})")

    if dsl.logic == "AND":
        combined = "*".join(cond_parts)
    else:
        combined = "SIGN(" + "+".join(cond_parts) + ")"

    _log.debug("    Combined condition expression: %s", combined)

    first_col = all_cols[0]
    row_range = src.range(first_col)
    anchor = f"INDIRECT(\"'{src.sheet_name}'!{src.col_letter(first_col)}$2\")"

    for out_col_idx, col_name in enumerate(all_cols, start=1):
        col_range = src.range(col_name)
        for data_row in range(1, src.data_rows + 1):
            excel_row = data_row + 1
            formula = (
                f"=IFERROR(INDEX({col_range},"
                f"AGGREGATE(15,6,"
                f"(ROW({row_range})-ROW({anchor})+1)"
                f"/({combined})"
                f",{data_row})),\"\")"
            )
            ws.cell(row=excel_row, column=out_col_idx, value=formula)

    _log.debug("    Wrote %d x %d filter formula cells", src.data_rows, len(all_cols))


def _write_groupby(
    dsl: GroupByAction,
    src: _SourceInfo,
    ws: Worksheet,
    extra_filters: list[FilterAction] | None = None,
) -> None:
    extra_filters = extra_filters or []
    _log.debug(
        "  [groupby] rows=%s  values=%s  extra_filters=%d",
        dsl.rows,
        [(v.aggregation, v.column) for v in dsl.values],
        len(extra_filters),
    )
    assert src.data_ws is not None, "groupby requires a source worksheet"

    # When filters are folded in, scan only rows that pass the filter.
    row_checks = _build_row_checks(extra_filters, src.data_col_indices) if extra_filters else []
    unique_keys = _unique_values(
        src.data_ws,
        [src.data_col_indices[r] for r in dsl.rows],
        row_checks,
    )
    _log.debug("    Found %d unique key combinations", len(unique_keys))
    if unique_keys:
        _log.debug("    First 5 keys: %s", unique_keys[:5])

    # Build extra SUMIFS/COUNTIFS criteria pairs from the accumulated filters.
    extra_criteria = _build_sumifs_criteria(extra_filters, src) if extra_filters else []
    if extra_criteria:
        _log.debug("    Extra SUMIFS criteria: %s", extra_criteria)

    headers: list[str] = list(dsl.rows)
    for v in dsl.values:
        headers.append(f"{v.aggregation}_{v.column}")
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=h)

    dim_count = len(dsl.rows)

    for row_offset, key_tuple in enumerate(unique_keys):
        excel_row = row_offset + 2
        for d, val in enumerate(key_tuple):
            ws.cell(row=excel_row, column=d + 1, value=val)

        for v_idx, agg in enumerate(dsl.values):
            out_col = dim_count + v_idx + 1
            formula = _agg_formula(agg, dsl.rows, key_tuple, src, extra_criteria)
            ws.cell(row=excel_row, column=out_col, value=formula)

    if unique_keys:
        sample_row = 2
        sample_col = dim_count + 1
        _log.debug("    Sample formula: %s", ws.cell(row=sample_row, column=sample_col).value)


def _write_pivot(
    dsl: PivotAction,
    src: _SourceInfo,
    ws: Worksheet,
    extra_filters: list[FilterAction] | None = None,
) -> None:
    extra_filters = extra_filters or []
    row_dims = dsl.rows
    col_dims = dsl.columns
    _log.debug(
        "  [pivot] row_dims=%s  col_dims=%s  values=%s  extra_filters=%d",
        row_dims,
        col_dims,
        [(v.aggregation, v.column) for v in dsl.values],
        len(extra_filters),
    )

    assert src.data_ws is not None, "pivot requires a source worksheet"
    row_checks = _build_row_checks(extra_filters, src.data_col_indices) if extra_filters else []
    row_keys = _unique_values(
        src.data_ws,
        [src.data_col_indices[r] for r in row_dims],
        row_checks,
    )
    col_keys = _unique_values(
        src.data_ws,
        [src.data_col_indices[c] for c in col_dims],
        row_checks,
    )
    _log.debug("    Row keys: %d  Col keys: %d", len(row_keys), len(col_keys))

    extra_criteria = _build_sumifs_criteria(extra_filters, src) if extra_filters else []

    header_offset = len(row_dims)
    for d, name in enumerate(row_dims):
        ws.cell(row=1, column=d + 1, value=name)

    for ck_idx, ck in enumerate(col_keys):
        for v_idx, agg in enumerate(dsl.values):
            out_col = header_offset + ck_idx * len(dsl.values) + v_idx + 1
            label = " / ".join(str(x) for x in ck) + f" ({agg.aggregation}_{agg.column})"
            ws.cell(row=1, column=out_col, value=label)

    for rk_idx, rk in enumerate(row_keys):
        excel_row = rk_idx + 2
        for d, val in enumerate(rk):
            ws.cell(row=excel_row, column=d + 1, value=val)

        for ck_idx, ck in enumerate(col_keys):
            for v_idx, agg in enumerate(dsl.values):
                out_col = header_offset + ck_idx * len(dsl.values) + v_idx + 1
                formula = _pivot_formula(agg, row_dims, col_dims, rk, ck, src, extra_criteria)
                ws.cell(row=excel_row, column=out_col, value=formula)

    total_cells = len(row_keys) * len(col_keys) * len(dsl.values)
    _log.debug("    Wrote %d pivot formula cells (%d rows x %d cols)", total_cells, len(row_keys), len(col_keys))


def _write_sort(dsl: SortAction, src: _SourceInfo, ws: Worksheet) -> None:
    _log.debug(
        "  [sort] orderBy=%s  direction=%s  limit=%s",
        dsl.order_by, dsl.direction, dsl.limit,
    )
    all_cols = src.columns
    for out_col_idx, col_name in enumerate(all_cols, start=1):
        ws.cell(row=1, column=out_col_idx, value=col_name)

    order_range = src.range(dsl.order_by)
    n_rows = dsl.limit if dsl.limit else src.data_rows

    first_col = all_cols[0]
    row_range = src.range(first_col)

    rank_fn = "SMALL" if dsl.direction == "ASC" else "LARGE"
    _log.debug("    rank_fn=%s  output_rows=%d", rank_fn, n_rows)

    for data_row in range(1, n_rows + 1):
        excel_row = data_row + 1
        for out_col_idx, col_name in enumerate(all_cols, start=1):
            col_range = src.range(col_name)
            formula = (
                f"=IFERROR(INDEX({col_range},"
                f"MATCH({rank_fn}({order_range},{data_row}),"
                f'{order_range},0)),"")'
            )
            ws.cell(row=excel_row, column=out_col_idx, value=formula)

    if n_rows > 0:
        _log.debug("    Sample formula [row 2]: %s", ws.cell(row=2, column=1).value)


# ---------------------------------------------------------------------------
# Formula helpers
# ---------------------------------------------------------------------------

_AGG_FN = {
    "SUM": "SUMIFS",
    "AVERAGE": "AVERAGEIFS",
    "COUNT": "COUNTIFS",
    "MAX": "MAXIFS",
    "MIN": "MINIFS",
}


def _formula_literal(v: object) -> str:
    """Convert a Python value to an Excel formula literal."""
    if isinstance(v, str):
        return f'"{v}"'
    if isinstance(v, datetime.datetime):
        return f'DATEVALUE("{v:%Y-%m-%d}")+TIMEVALUE("{v:%H:%M:%S}")'
    if isinstance(v, datetime.date):
        return f'DATEVALUE("{v:%Y-%m-%d}")'
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return str(v)


def _agg_formula(
    agg: AggregationValue,
    dim_names: list[str],
    key_values: tuple,
    src: _SourceInfo,
    extra_criteria: list[str] | None = None,
) -> str:
    fn = _AGG_FN[agg.aggregation]
    val_range = src.range(agg.column)
    extra = extra_criteria or []

    if fn == "COUNTIFS":
        parts = []
        for dim, val in zip(dim_names, key_values):
            parts.append(f"{src.range(dim)},{_formula_literal(val)}")
        parts.extend(extra)
        return f"={fn}({','.join(parts)})"

    parts = [val_range]
    for dim, val in zip(dim_names, key_values):
        parts.append(f"{src.range(dim)},{_formula_literal(val)}")
    parts.extend(extra)
    formula_body = f"{fn}({','.join(parts)})"
    if extra:
        return f"=IFERROR({formula_body},\"\")"
    return f"={formula_body}"


def _pivot_formula(
    agg: AggregationValue,
    row_dims: list[str],
    col_dims: list[str],
    row_key: tuple,
    col_key: tuple,
    src: _SourceInfo,
    extra_criteria: list[str] | None = None,
) -> str:
    fn = _AGG_FN[agg.aggregation]
    val_range = src.range(agg.column)
    extra = extra_criteria or []

    criteria_parts: list[str] = []
    for dim, val in list(zip(row_dims, row_key)) + list(zip(col_dims, col_key)):
        crit_range = src.range(dim)
        criteria_parts.append(f"{crit_range},{_formula_literal(val)}")
    criteria_parts.extend(extra)

    if fn == "COUNTIFS":
        return f"={fn}({','.join(criteria_parts)})"

    formula_body = f"{fn}({val_range},{','.join(criteria_parts)})"
    if extra:
        return f"=IFERROR({formula_body},\"\")"
    return f"={formula_body}"


def _unique_values(
    ws: Worksheet,
    col_indices: list[int],
    row_checks: list[tuple[int, str, object]] | None = None,
) -> list[tuple]:
    """Extract unique tuples from the given column indices (1-based).

    When *row_checks* is provided, only rows passing all checks are included.
    """
    seen: set[tuple] = set()
    result: list[tuple] = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row_checks and not _row_passes(row, row_checks):
            continue
        key = tuple(row[i - 1].value for i in col_indices)
        if key not in seen and any(v is not None for v in key):
            seen.add(key)
            result.append(key)
    result.sort(key=lambda t: tuple(str(x) for x in t))
    return result
