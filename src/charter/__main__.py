"""CLI entry-point: ``uv run charter input.xlsx [-o output.xlsx]``."""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import click
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

from . import log
from .analyzer import analyze
from .config import init_config, load_config
from .dsl import AnalysisDimension
from .executor import execute_dsl
from .reader import read_schema, schema_to_prompt

console = Console()


@click.command()
@click.argument("input_file", type=click.Path(exists=True, dir_okay=False))
@click.option(
    "-o", "--output",
    type=click.Path(dir_okay=False),
    default=None,
    help="Output Excel file path.  Defaults to <input>_analysis.xlsx.",
)
@click.option(
    "-c", "--config",
    "config_path",
    type=click.Path(dir_okay=False),
    default=None,
    help="Path to charter.toml config file.  Auto-detected from cwd or ~/.config/charter/ if omitted.",
)
@click.option(
    "-v", "--verbose",
    is_flag=True,
    default=False,
    help="Enable verbose debug output.",
)
@click.option(
    "--init-config",
    "do_init_config",
    is_flag=True,
    default=False,
    help="Write a default charter.toml to the current directory and exit.",
)
def main(
    input_file: str,
    output: str | None,
    config_path: str | None,
    verbose: bool,
    do_init_config: bool,
) -> None:
    """Analyse an Excel file with AI and generate formula-driven insight sheets."""

    if do_init_config:
        dest = init_config()
        console.print(f"[green]Default config written to[/green] [underline]{dest}[/underline]")
        console.print("Edit it to set your API key, base URL, model, and prompt.")
        return

    # --- 0. Load config ---
    try:
        cfg = load_config(config_path)
    except FileNotFoundError as exc:
        console.print(f"[bold red]Config error:[/] {exc}")
        sys.exit(1)

    # --verbose flag overrides config
    debug = verbose or cfg.debug
    log.setup(verbose=debug)
    _log = log.get("cli")

    _log.debug("Config loaded  debug=%s  model=%s  base_url=%s", debug, cfg.model, cfg.base_url or "(default)")

    console.print(
        f"[dim]Config: model={cfg.model}  base_url={cfg.base_url or '(default)'}  debug={debug}[/dim]"
    )

    input_path = Path(input_file)
    if output is None:
        output_path = input_path.with_stem(input_path.stem + "_analysis")
    else:
        output_path = Path(output)

    _log.debug("Input: %s", input_path.resolve())
    _log.debug("Output: %s", output_path.resolve())

    # --- 1. Read schema ---
    with console.status("[bold green]Reading Excel file..."):
        schema, ws_source = read_schema(input_path)

    console.print(Panel(schema_to_prompt(schema), title="Data Schema", border_style="cyan"))

    # --- 2. Call AI ---
    with console.status(f"[bold yellow]Calling {cfg.model} for analysis recommendations..."):
        try:
            response = analyze(schema, cfg=cfg)
        except Exception as exc:
            _log.debug("AI analysis failed", exc_info=True)
            console.print(f"[bold red]AI analysis failed:[/] {exc}")
            sys.exit(1)

    _display_dimensions(response.dimensions)

    # --- 3. Display generated DSL for each dimension in debug mode ---
    if debug:
        for dim in response.dimensions:
            dsl_json = dim.dsl.model_dump_json(indent=2, by_alias=True)
            console.print(
                Panel(dsl_json, title=f"DSL: {dim.title}", border_style="yellow")
            )

    # --- 4. Let user confirm ---
    if not click.confirm("\nProceed to generate analysis sheets?", default=True):
        console.print("[dim]Aborted.[/dim]")
        return

    # --- 5. Copy workbook, execute DSL, save ---
    shutil.copy2(input_path, output_path)
    _log.debug("Copied %s -> %s", input_path, output_path)

    from openpyxl import load_workbook

    wb = load_workbook(str(output_path))
    ws_data = wb.active

    with console.status("[bold green]Generating formula sheets..."):
        for dim in response.dimensions:
            try:
                execute_dsl(dim.dsl, schema, ws_data, wb, dim.title)
                console.print(f"  [green]\u2713[/green] {dim.title}")
            except Exception as exc:
                _log.debug("Failed to execute DSL for '%s'", dim.title, exc_info=True)
                console.print(f"  [red]\u2717[/red] {dim.title}: {exc}")

    wb.save(str(output_path))
    _log.debug("Workbook saved: %d sheets total", len(wb.sheetnames))

    console.print(f"\n[bold green]Done![/bold green] Output saved to [underline]{output_path}[/underline]")


def _display_dimensions(dims: list[AnalysisDimension]) -> None:
    table = Table(title="Recommended Analysis Dimensions", show_lines=True)
    table.add_column("#", style="bold", width=3)
    table.add_column("Title", style="cyan")
    table.add_column("Type", style="magenta")
    table.add_column("Description")

    for i, dim in enumerate(dims, start=1):
        actions = " -> ".join(s.action for s in dim.dsl.pipeline)
        table.add_row(str(i), dim.title, actions, dim.description)

    console.print(table)


if __name__ == "__main__":
    main()
